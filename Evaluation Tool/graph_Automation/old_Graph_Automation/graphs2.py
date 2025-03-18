import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import FuncFormatter

#sheet_name = "Encryptors"
#sheet_name = "message_Size"
sheet_name = "Looping"
#sheet_name = "parallel_Opening"
#sheet_name = "parallel_Opening_Closing"
#sheet_name = "exclusive_Opening"
#sheet_name = "exclusive_Opening_Closing"

file_name = "performance_Analysis.xlsx"
file_name = "cost_Analysis.xlsx"


def format_number(num, file_name):
    if file_name == "cost_Analysis.xlsx":
        if num >= 1_000:
            return f"{num/1_000:,.0f}K"  # Format as thousands with 'K'
        else:
            return f"{num:,.0f}"  # Return the number as it is if it's less than 1,000
    else:
        return f"{num:,.0f}"

def no_scientific(x, pos):
    if file_name == "cost_Analysis.xlsx":
        if x >= 1e3:  # Only format values greater than or equal to 1000
            return f'{x / 1e3:,.0f}K'  # Format as thousands with 'K'
        else:
            return f'{x:,.0f}'  # For smaller values, display as is
    else:
        return f'{x:,.0f}'
    

workbook = openpyxl.load_workbook(file_name, data_only=True)
sheet = workbook[sheet_name]

# Plotting stacked bars
fig, ax = plt.subplots(figsize=(10, 6))

if file_name == "performance_Analysis.xlsx":
    ax.set_ylabel('Cumulative exec. time [ms]', fontsize=12)
    
    if sheet_name == "Encryptors" or sheet_name == "Looping":
        res = []
        x_labels = []
        for table_name in sheet.tables:
            table = sheet.tables[table_name]
            if sheet_name == "Encryptors":
                x_labels.append(table_name.split("_")[0])
                plt.xlabel('Number of encryptors', fontsize=12)
            elif sheet_name == "Looping":
                x_labels.append(table_name.split("_")[1])
                plt.xlabel('Loops', fontsize=12)
            header_values = [cell.value for cell in sheet[table.ref][0]]
            avg_index = header_values.index("Average")
            functions_index = header_values.index("Function/Endpoint")
            type_index = header_values.index("Type")
            summo = 0
            res2 = []
            done = False
            for row in sheet[table.ref][1:]:
                if row[type_index].value == "Rest":
                    summo += row[avg_index].value
                    if row[functions_index].value == "saveModel":
                        res2.append(int(summo))
                        summo = 0
                    if row[functions_index].value == "attributesCertification":
                        res2.append(int(summo))
                        summo = 0
                    elif "decrypt" in row[functions_index].value and done == False:
                        res2.append(int(summo - row[avg_index].value)) 
                        done = True
                        summo = row[avg_index].value
            res2.append(int(summo))
            res.append(res2)

    else:
        res = []
        x_labels = []
        for table_name in sheet.tables:
            table = sheet.tables[table_name]
            if sheet_name == "message_Size":
                x_labels.append(table_name.split("_")[1])
                plt.xlabel('Size duplication', fontsize=12)
            elif sheet_name == "parallel_Opening":
                x_labels.append(table_name.split("_")[1])
                plt.xlabel('Openings', fontsize=12)
            elif sheet_name == "exclusive_Opening":
                x_labels.append(table_name.split("_")[1])
                plt.xlabel('Openings', fontsize=12)
            elif sheet_name == "parallel_Opening_Closing":
                x_labels.append(table_name.split("_")[2])
                plt.xlabel('Openings and Closings', fontsize=12)
            elif sheet_name == "exclusive_Opening_Closing":
                x_labels.append(table_name.split("_")[2])
                plt.xlabel('Openings and Closings', fontsize=12)
            header_values = [cell.value for cell in sheet[table.ref][0]]
            avg_index = header_values.index("Average")
            
            summo = 0
            res2 = []
            row_new = 2
            done = False
            for row in sheet[table.ref][1:]:
                type_index = sheet.cell(row = row_new, column = 1).value
                functions_index = sheet.cell(row = row_new, column = 2).value
                if type_index == "Rest":
                    summo += row[avg_index].value
                    if functions_index == "saveModel":
                        res2.append(int(summo))
                        summo = 0
                    if functions_index == "attributesCertification":
                        res2.append(int(summo))
                        summo = 0
                    elif "decrypt" in functions_index and done == False:
                        res2.append(int(summo - row[avg_index].value)) 
                        done = True
                        summo = row[avg_index].value
                row_new += 1
            res2.append(int(summo))
            res.append(res2)
    

else:
    ax.set_ylabel("Cumulative gas used", fontsize=12)
    res = []
    x_labels = []
    for table_name in sheet.tables:
        if sheet_name == "Encryptors":
            x_labels.append(table_name.split("_")[0])
            plt.xlabel('Number of encryptors', fontsize=12)
        elif sheet_name == "Looping":
            x_labels.append(table_name.split("_")[1])
            plt.xlabel('Loops', fontsize=12)
        elif sheet_name == "message_Size":
            x_labels.append(table_name.split("_")[1])
            plt.xlabel('Size duplication', fontsize=12)
        elif sheet_name == "parallel_Opening":
            x_labels.append(table_name.split("_")[1])
            plt.xlabel('Openings', fontsize=12)
        elif sheet_name == "exclusive_Opening":
            x_labels.append(table_name.split("_")[1])
            plt.xlabel('Openings', fontsize=12)
        elif sheet_name == "parallel_Opening_Closing":
            x_labels.append(table_name.split("_")[2])
            plt.xlabel('Openings and Closings', fontsize=12)
        elif sheet_name == "exclusive_Opening_Closing":
            x_labels.append(table_name.split("_")[2])
            plt.xlabel('Openings and Closings', fontsize=12)
        summo = 0
        res2 = []
        done = False
        done2 = False
        table = sheet.tables[table_name]
        header_values = [cell.value for cell in sheet[table.ref][0]]
        gas_index = header_values.index("Gas")
        functions_index = header_values.index("Function")
        for row in sheet[table.ref][1:]:
            if row[functions_index].value == "setIPFSLink" and done == False:
                res2.append(summo)
                done = True
                summo = 0
            elif row[functions_index].value == "notifyAuthorities" and done2 == False:
                res2.append(summo)
                done2 = True
                summo = 0
            summo += row[gas_index].value
        res2.append(summo)
        res.append(res2)
        
        
data = np.array(res)
num_points = data.shape[0]
x = np.arange(num_points)

# Bar width
bar_width = 0.6

# Colors with transparency
colors = ['#6fa3f7', '#ff9e6a', '#8fd34e', '#ff77b6']  # Blue, Orange, Green, Pink
alpha_val = 0.8  # Transparency level

# Plot the stacked bars with black outline
if file_name != "performance_Analysis.xlsx":
    ax.bar(x, data[:, 0], bar_width, color=colors[0], label='Instantiate', alpha=alpha_val, edgecolor='gray', linewidth=1)  # Phase 1
    ax.bar(x, data[:, 1], bar_width, bottom=data[:, 0], color=colors[1], label='Transact', alpha=alpha_val, edgecolor='gray', linewidth=1)  # Phase 2
    ax.bar(x, data[:, 2], bar_width, bottom=data[:, 0] + data[:, 1], color=colors[2], label='Inspect', alpha=alpha_val, edgecolor='gray', linewidth=1)  # Phase 3
else:
    ax.bar(x, data[:, 0], bar_width, color=colors[0], label='Configure', alpha=alpha_val, edgecolor='gray', linewidth=1)  # Phase 1
    ax.bar(x, data[:, 1], bar_width, bottom=data[:, 0], color=colors[1], label='Instantiate', alpha=alpha_val, edgecolor='gray', linewidth=1)  # Phase 2
    ax.bar(x, data[:, 2], bar_width, bottom=data[:, 0] + data[:, 1], color=colors[2], label='Transact', alpha=alpha_val, edgecolor='gray', linewidth=1)  # Phase 3
    ax.bar(x, data[:, 3], bar_width, bottom=data[:, 0] + data[:, 1] + data[:, 2], color=colors[3], label='Inspect', alpha=alpha_val, edgecolor='gray', linewidth=1)

# Adding labels on top of each segment
for i in range(num_points):
    ax.text(x[i], data[i, 0]/2, format_number(data[i, 0], file_name), ha='center', va='center', color='black')
    ax.text(x[i], data[i, 0] + data[i, 1]/2, format_number(data[i, 1], file_name), ha='center', va='center', color='black')
    ax.text(x[i], data[i, 0] + data[i, 1] + data[i, 2]/2, format_number(data[i, 2], file_name), ha='center', va='center', color='black')
    if file_name == "performance_Analysis.xlsx":
        ax.text(x[i], data[i, 0] + data[i, 1] + data[i, 2] + data[i, 3]/2, format_number(data[i, 3], file_name), ha='center', va='center', color='black')

# Labels and Formatting
ax.set_xticks(x)
ax.set_xticklabels(x_labels, fontsize=10)

# Adding custom y-ticks with equally spaced values
max_height = np.max(np.sum(data, axis=1))  # Maximum height of the stacked bars
y_ticks = np.linspace(0, max_height, 7)  # 7 equally spaced values from 0 to max_height
ax.set_yticks(y_ticks)


# Format the y-tick labels with the same formatting
ax.yaxis.set_major_formatter(FuncFormatter(no_scientific))
# Adding gridlines for better readability
ax.grid(axis='y', linestyle='--', alpha=0.7)

# Add legend
ax.legend(title="Phases", loc='center', fontsize=10, bbox_to_anchor=(0.5, -0.15), ncol=4)

# Show plot
plt.tight_layout()
plt.show()
