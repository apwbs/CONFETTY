import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import FuncFormatter

sheet_name = "exclusive_Opening"
sheet_name2 = "exclusive_Opening_Closing"


file_name = "performance_Analysis.xlsx"
file_name = "cost_Analysis.xlsx"

def format_number(num, file_name):
    if file_name == "cost_Analysis.xlsx":
        if num >= 1_000:
            return f"{num/1_000:,.0f}K"  # Format as thousands with 'K'
        else:
            return f"{num:,.0f}"  # Return the number as it is if it's less than 1,000
    else:
        return f"{num:,.0f}"

def no_scientific(x, pos):
    if file_name == "cost_Analysis.xlsx":
        if x >= 1e3:  # Only format values greater than or equal to 1000
            return f'{x / 1e3:,.0f}K'  # Format as thousands with 'K'
        else:
            return f'{x:,.0f}'  # For smaller values, display as is
    else:
        return f'{x:,.0f}'
    
workbook = openpyxl.load_workbook(file_name, data_only=True)
sheet_1 = workbook[sheet_name]
sheet_2 = workbook[sheet_name2]

if file_name == "performance_Analysis.xlsx":
    res_1 = []
    x_labels_1 = []
    for table_name in sheet_1.tables:
        table = sheet_1.tables[table_name]
        x_labels_1.append(table_name.split("_")[1])
        plt.xlabel('Openings', fontsize=12)
        
        header_values = [cell.value for cell in sheet_1[table.ref][0]]
        avg_index = header_values.index("Average")
        summo = 0
        res2 = []
        row_new = 2
        done = False
        for row in sheet_1[table.ref][1:]:
            type_index = sheet_1.cell(row = row_new, column = 1).value
            functions_index = sheet_1.cell(row = row_new, column = 2).value
            if type_index == "Rest":
                summo += row[avg_index].value
                if functions_index == "saveModel":
                    res2.append(int(summo))
                    summo = 0
                if functions_index == "attributesCertification":
                    res2.append(int(summo))
                    summo = 0
                elif "decrypt" in functions_index and done == False:
                    res2.append(int(summo - row[avg_index].value)) 
                    done = True
                    summo = row[avg_index].value
            row_new += 1
        res2.append(int(summo))
        res_1.append(res2)
        
        
    res_2 = []
    x_labels_2 = []
    for table_name in sheet_2.tables:
        table = sheet_2.tables[table_name]
        x_labels_2.append(table_name.split("_")[2])
        plt.xlabel('Openings and Closings', fontsize=12)
        
        header_values = [cell.value for cell in sheet_2[table.ref][0]]
        avg_index = header_values.index("Average")
        summo = 0
        res2 = []
        row_new = 2
        done = False
        for row in sheet_2[table.ref][1:]:
            type_index = sheet_2.cell(row = row_new, column = 1).value
            functions_index = sheet_2.cell(row = row_new, column = 2).value
            if type_index == "Rest":
                summo += row[avg_index].value
                if functions_index == "saveModel":
                    res2.append(int(summo))
                    summo = 0
                if functions_index == "attributesCertification":
                    res2.append(int(summo))
                    summo = 0
                elif "decrypt" in functions_index and done == False:
                    res2.append(int(summo - row[avg_index].value)) 
                    done = True
                    summo = row[avg_index].value
            row_new += 1
        res2.append(int(summo))
        res_2.append(res2)
        
else:
    res_1 = []
    x_labels_1 = []
    for table_name in sheet_1.tables:
        x_labels_1.append(table_name.split("_")[1])
        plt.xlabel('Openings', fontsize=12)
        summo = 0
        res2 = []
        done = False
        done2 = False
        table = sheet_1.tables[table_name]
        header_values = [cell.value for cell in sheet_1[table.ref][0]]
        gas_index = header_values.index("Gas")
        functions_index = header_values.index("Function")
        for row in sheet_1[table.ref][1:]:
            if row[functions_index].value == "setIPFSLink" and done == False:
                res2.append(int(summo))
                done = True
                summo = 0
            elif row[functions_index].value == "notifyAuthorities" and done2 == False:
                res2.append(int(summo))
                done2 = True
                summo = 0
            summo += row[gas_index].value
        res2.append(int(summo))
        res_1.append(res2)
        
    
    res_2 = []
    x_labels_2 = []
    for table_name in sheet_2.tables:
        x_labels_2.append(table_name.split("_")[2])
        plt.xlabel('Openings and Closings', fontsize=12)
        
        summo = 0
        res2 = []
        done = False
        done2 = False
        table = sheet_2.tables[table_name]
        header_values = [cell.value for cell in sheet_2[table.ref][0]]
        gas_index = header_values.index("Gas")
        functions_index = header_values.index("Function")
        for row in sheet_2[table.ref][1:]:
            if row[functions_index].value == "setIPFSLink" and done == False:
                res2.append(int(summo))
                done = True
                summo = 0
            elif row[functions_index].value == "notifyAuthorities" and done2 == False:
                res2.append(int(summo))
                done2 = True
                summo = 0
            summo += row[gas_index].value
        res2.append(int(summo))
        res_2.append(res2)


# Generate X-axis labels
x_labels = [f"x{e+1}" for e in range(len(x_labels_1))]

width = 0.3  # Bar width
gap = 0.08   # Gap between bars

# Define colors and labels
colors = ['#6fa3f7', '#ff9e6a', '#8fd34e', '#ff77b6']  # Blue, Orange, Green, Pink
if file_name == "performance_Analysis.xlsx":
    labels = ['Configure', 'Instantiate', 'Transact', 'Inspect']  # Legend labels
else:
    labels = ['Instantiate', 'Transact', 'Inspect']  # Legend labels

# Numeric X-axis positions
x_numeric = np.arange(len(x_labels))
x_res1 = x_numeric - (width / 2 + gap / 2)
x_res2 = x_numeric + (width / 2 + gap / 2)

fig, ax = plt.subplots(figsize=(10, 6))

# Stack levels
bottom_1 = np.zeros(len(res_1))
bottom_2 = np.zeros(len(res_2))

# Compute max bar height
max_height_1 = np.max(np.sum(res_1, axis=1))  # Highest sum in res_1
max_height_2 = np.max(np.sum(res_2, axis=1))  # Highest sum in res_2
max_height = max(max_height_1, max_height_2)  # Overall max

# Create equally spaced Y-axis ticks (max + 6 values down to 0)
yticks = np.linspace(0, max_height, 7)

# Plot stacked bars for res_1
for i in range(len(labels)):
    values = [row[i] for row in res_1]
    bars = ax.bar(x_res1, values, width, bottom=bottom_1, color=colors[i], edgecolor='gray', label=labels[i], alpha=0.8)

    # Add text on each segment
    for bar, val in zip(bars, values):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_y() + bar.get_height() / 2, 
                format_number(val, file_name), ha='center', va='center', fontsize=5, color='black')

    bottom_1 += values

# Plot stacked bars for res_2
for i in range(len(labels)):
    values = [row[i] for row in res_2]
    bars = ax.bar(x_res2, values, width, bottom=bottom_2, color=colors[i], edgecolor='gray', alpha=0.8)

    # Add text on each segment
    for bar, val in zip(bars, values):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_y() + bar.get_height() / 2, 
                format_number(val, file_name), ha='center', va='center', fontsize=5, color='black')

    bottom_2 += values

# Labels and legend
ax.set_xlabel('Exclusive opening / Exclusive opening and closing', fontsize=12)
ax.set_xticks(x_numeric)
ax.set_xticklabels(x_labels)

# Set Y-axis ticks
ax.set_yticks(yticks)
ax.yaxis.set_major_formatter(FuncFormatter(no_scientific))

# Add Y-axis grid lines (gray, dashed)
ax.yaxis.grid(True, linestyle='--', linewidth=0.7, alpha=0.6)  

# Move grid lines to the background
ax.set_axisbelow(True)

if file_name == "performance_Analysis.xlsx":
    ax.set_ylabel('Cumulative exec. time [ms]', fontsize=12)
else:
    ax.set_ylabel("Cumulative gas used", fontsize=12)

ax.legend(title="Phases", loc='center', fontsize=10, bbox_to_anchor=(0.5, -0.16), ncol=4)

plt.show()