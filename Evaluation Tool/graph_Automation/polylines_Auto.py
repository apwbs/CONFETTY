import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import FuncFormatter
import os

# ==========================
# SHEET AND FILE SELECTION
# ==========================

sheet_name = "writing_Participants"
#sheet_name = "message_Size"
#sheet_name = "process_Size"
#sheet_name = "parallel_Split"
#sheet_name = "parallel_Split_Join"
#sheet_name = "exclusive_Split"
#sheet_name = "exclusive_Split_Join"

file_name = "performance_Analysis.xlsx"
file_name = "cost_Analysis.xlsx"

# ==========================
# FORMATTER FUNCTION
# ==========================

def no_scientific(x, pos):
    """Custom formatter for Y-axis to avoid scientific notation and show commas."""
    return f'{x:,.0f}'

def roman_to_int_small(s):
    # Mapping only the numerals from 2 to 10
    roman_map = {
        'II': 2,
        'III': 3,
        'IV': 4,
        'V': 5,
        'VI': 6,
        'VII': 7,
        'VIII': 8,
        'IX': 9,
        'X': 10
    }

    # Return the integer value or None if not found
    return roman_map.get(s, None)

# ==========================
# LOAD WORKBOOK AND SHEET
# ==========================

workbook = openpyxl.load_workbook(file_name, data_only=True)
sheet = workbook[sheet_name]

# ==========================
# DATA COLLECTION
# ==========================

x_labels = []  # X-axis labels for the plot
res = []       # Collected data results

if file_name == "performance_Analysis.xlsx":

    # ======================
    # CASE: performance_Analysis
    # ======================

    if sheet_name in ["writing_Participants", "process_Size"]:
        # Process sheets focused on writing_Participants or process_Size cases
        for table_name in sheet.tables:
            table = sheet.tables[table_name]

            # Extract X-axis label based on sheet type
            if sheet_name == "writing_Participants":
                roman = table_name.split("_")[0]
                x_labels.append(roman_to_int_small(roman))
            elif sheet_name == "process_Size":
                x_labels.append(table_name.split("_")[1])

            # Extract header indexes
            header_values = [cell.value for cell in sheet[table.ref][0]]
            avg_index = header_values.index("Average")
            functions_index = header_values.index("Function/Endpoint")
            type_index = header_values.index("Type")

            # Iterate through table rows
            summo = 0
            res2 = []
            done = False

            for row in sheet[table.ref][1:]:
                if row[type_index].value == "Rest":
                    summo += row[avg_index].value

                    if row[functions_index].value == "saveModel":
                        res2.append(int(summo))
                        summo = 0

                    if row[functions_index].value == "attributesCertification":
                        res2.append(int(summo))
                        summo = 0

                    elif "decrypt" in row[functions_index].value and not done:
                        res2.append(int(summo - row[avg_index].value))
                        done = True
                        summo = row[avg_index].value

            res2.append(int(summo))  # Append the remaining value
            res.append(res2)

    else:
        # ======================
        # OTHER SHEET NAMES (message_Size, parallel/exclusive openings)
        # ======================

        for table_name in sheet.tables:
            table = sheet.tables[table_name]

            # Extract X-axis label based on sheet type
            if sheet_name == "message_Size":
                x_labels.append(table_name.split("_")[1])
            elif sheet_name in ["parallel_Split", "exclusive_Split"]:
                x_labels.append(table_name.split("_")[1])
            elif sheet_name in ["parallel_Split_Join", "exclusive_Split_Join"]:
                x_labels.append(table_name.split("_")[2])

            # Extract header indexes
            header_values = [cell.value for cell in sheet[table.ref][0]]
            avg_index = header_values.index("Average")

            summo = 0
            res2 = []
            done = False
            row_new = 2  # Excel is 1-indexed; header is row 1

            # Iterate through table rows
            for row in sheet[table.ref][1:]:
                type_index = sheet.cell(row=row_new, column=1).value
                functions_index = sheet.cell(row=row_new, column=2).value

                if type_index == "Rest":
                    summo += row[avg_index].value

                    if functions_index == "saveModel":
                        res2.append(int(summo))
                        summo = 0

                    if functions_index == "attributesCertification":
                        res2.append(int(summo))
                        summo = 0

                    elif "decrypt" in functions_index and not done:
                        res2.append(int(summo - row[avg_index].value))
                        done = True
                        summo = row[avg_index].value

                row_new += 1

            res2.append(int(summo))  # Append the remaining value
            res.append(res2)

else:
    # ======================
    # CASE: cost_Analysis
    # ======================

    for table_name in sheet.tables:
        table = sheet.tables[table_name]

        # Extract X-axis label based on sheet type
        if sheet_name == "writing_Participants":
            roman = table_name.split("_")[0]
            x_labels.append(roman_to_int_small(roman))
        elif sheet_name == "process_Size":
            x_labels.append(table_name.split("_")[1])
        elif sheet_name == "message_Size":
            x_labels.append(table_name.split("_")[1])
        elif sheet_name in ["parallel_Split", "exclusive_Split"]:
            x_labels.append(table_name.split("_")[1])
        elif sheet_name in ["parallel_Split_Join", "exclusive_Split_Join"]:
            x_labels.append(table_name.split("_")[2])

        # Extract header indexes
        header_values = [cell.value for cell in sheet[table.ref][0]]
        gas_index = header_values.index("Gas")
        functions_index = header_values.index("Function")

        summo = 0
        res2 = []
        done = False
        done2 = False

        # Iterate through table rows
        for row in sheet[table.ref][1:]:
            func = row[functions_index].value

            if func == "setIPFSLink" and not done:
                res2.append(summo)
                done = True
                summo = 0

            elif func == "notifyAuthorities" and not done2:
                res2.append(summo)
                done2 = True
                summo = 0

            summo += row[gas_index].value

        res2.append(summo)  # Append the remaining gas value
        res.append(res2)

# ==========================
# DATA VISUALIZATION
# ==========================

data = np.array(res)
num_points = data.shape[0]
x = np.arange(num_points)

# Compute cumulative sums for stacked area chart
cumulative_data = np.cumsum(data, axis=1)

# Create the plot
fig, ax = plt.subplots(figsize=(10, 6))

# Define color palette
fixed_colors = ['#ef476f', '#ffd166', '#06d6a0', '#118ab2', '#073b4c']
alpha_val = 0.8

# Set phase labels and colors based on the file type
if file_name == "performance_Analysis.xlsx":
    phase_labels = ['Configure', 'Instantiate', 'Transact', 'Inspect']
    colors = fixed_colors[:4]
else:
    phase_labels = ['Instantiate', 'Transact', 'Inspect']
    colors = fixed_colors[1:4]

# Plot cumulative lines and areas
previous = np.zeros(num_points)

for phase in range(cumulative_data.shape[1]):
    y = cumulative_data[:, phase]

    # Fill area between previous and current phase
    ax.fill_between(x, previous, y, color=colors[phase], alpha=0.3)

    # Plot line for the current phase
    ax.plot(x, y, marker='o', color=colors[phase], label=phase_labels[phase], alpha=alpha_val)

    previous = y

# ==========================
# AXIS CONFIGURATION
# ==========================

ax.set_xticks(x)
ax.set_xticklabels(x_labels)

# Set dynamic x-axis label based on the sheet_name
if sheet_name == "writing_Participants":
    xlabel = 'Number of writing participants'
    file_label = 'Number_Writing_Participants'
elif sheet_name == "process_Size":
    xlabel = 'Process size dimension'
    file_label = 'Process_Size'
elif sheet_name == "message_Size":
    xlabel = 'Message size dimension'
    file_label = 'Message_Size'
elif sheet_name == "parallel_Split":
    xlabel = 'Parallel splits'
    file_label = 'Parallel_Split'
elif sheet_name == "exclusive_Split":
    xlabel = 'Exclusive splits'
    file_label = 'Exclusive_Split'
elif sheet_name == "parallel_Split_Join":
    xlabel = 'Parallel splits and joins'
    file_label = 'Parallel_Split_Join'
elif sheet_name == "exclusive_Split_Join":
    xlabel = 'Exclusive splits and joins'
    file_label = 'Exclusive_Split_Join'

# Set the X-axis label (human-readable)
ax.set_xlabel(xlabel, fontsize=24)

# Set Y-axis label depending on the file type
if file_name == "performance_Analysis.xlsx":
    ax.set_ylabel("Cumulative exec. time [ms]", fontsize=24)
    tick_step = 2_000
    if file_label == 'Process_Size':
        tick_step = 10_000
    pdf_prefix = "time"
else:
    ax.set_ylabel("Cumulative gas used [GU]", fontsize=24)
    tick_step = 1_000_000
    if file_label == 'Process_Size':
        tick_step = 10_000_000
    pdf_prefix = "gas"

# Compute the min and max values for Y-axis
y_min = cumulative_data.min()
y_max = cumulative_data.max()

y_min_adjusted = (y_min // tick_step) * tick_step

how_much = 1

if file_name == "performance_Analysis.xlsx" and sheet_name =="writing_Participants":
    how_much = 3 
elif file_name == "performance_Analysis.xlsx" and sheet_name == "message_Size":
    how_much = 4
elif file_name == "performance_Analysis.xlsx" and (sheet_name =="exclusive_Split_Join" or sheet_name == "exclusive_Split" or sheet_name == "parallel_Split_Join" or sheet_name == "parallel_Split"):
    how_much = 2
elif file_name == "cost_Analysis.xlsx" and sheet_name == "message_Size":
    how_much = 2

y_max_adjusted = ((y_max // tick_step) + how_much) * tick_step

# Generate Y-axis ticks
y_ticks = np.arange(y_min_adjusted, y_max_adjusted + tick_step, tick_step)
ax.set_yticks(y_ticks)

# Set Y-axis limits to avoid cutting the lowest point
ax.set_ylim(bottom=y_min_adjusted)

# Apply Y-axis formatter to avoid scientific notation
ax.yaxis.set_major_formatter(FuncFormatter(no_scientific))

# Add grid, legend, and apply layout adjustments
ax.grid(axis='y', linestyle='--', alpha=0.7)
ax.tick_params(axis='x', labelsize=18)  # X-axis tick label size
ax.tick_params(axis='y', labelsize=18)  # Y-axis tick label size
ax.legend(title="Phases", loc='upper left', fontsize=16, title_fontsize=18)

plt.tight_layout()

# ==========================
# SAVE AS PDF
# ==========================

# Construct the filename using the prefix and file_label
pdf_filename = f"{pdf_prefix}_{file_label}.pdf"

# Optional: Choose a folder path to save the PDF
output_folder = "pdf_Outputs"
os.makedirs(output_folder, exist_ok=True)

# Complete path for the file
pdf_path = os.path.join(output_folder, pdf_filename)

# Save figure as PDF
plt.savefig(pdf_path, format='pdf')

print(f"Figure saved as: {pdf_path}")

# Show plot (optional)
plt.show()
